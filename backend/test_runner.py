"""
test_runner.py — сквозной (end-to-end) прогон API velo-system.

Запускается против ЖИВОГО сервера (не через ASGI-транспорт), поэтому
перед запуском поднимите backend:

    docker compose up --build
    # или локально:
    python -m uvicorn app.main:app --reload --host 127.0.0.1 --port 8000

Запуск:
    python test_runner.py

Переменные окружения (см. .env / .env.example):
    TEST_BASE_URL        — по умолчанию http://localhost:8000
    FIRST_ADMIN_EMAIL    — email первого admin-пользователя (создаётся init_db при первом старте)
    FIRST_ADMIN_PASSWORD — его пароль

Скрипт логинится под первым admin'ом, регистрирует временных
mechanic/manager пользователей для проверки ролей, прогоняет
все эндпоинты (включая права доступа и новый функционал 2.1-2.7),
и в конце удаляет все созданные им данные через cleanup().
"""

import os
import sys
import time
import uuid

import httpx
from dotenv import load_dotenv

# Нужен только для сверки чисел внутри выгруженного xlsx. Он есть в
# backend/requirements.txt, но test_runner.py могут запускать и из окружения,
# где его нет, — тогда пропускаем именно эту проверку, а не весь прогон.
try:
    from io import BytesIO
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

load_dotenv()

# Вывод содержит ✓/✗ — на Windows консоль по умолчанию использует
# кодировку вроде cp1251, которая их не знает, и print() падает.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

BASE_URL = os.getenv("TEST_BASE_URL", "http://localhost:8000")
ADMIN_EMAIL = os.getenv("FIRST_ADMIN_EMAIL", "velo@velo.local")
ADMIN_PASSWORD = os.getenv("FIRST_ADMIN_PASSWORD", "changeme")

RUN_ID = uuid.uuid4().hex[:8]  # чтобы не конфликтовать с данными от прошлых прогонов

client = httpx.Client(base_url=BASE_URL, timeout=15.0)

# Реестр созданных сущностей для cleanup(), в порядке, обратном созданию
created = {
    "repairs": [],
    "rentals": [],
    "person_tags": [],   # (person_id, tag_id)
    "passports": [],     # person_id
    "people": [],
    "bikes": [],
    "parts": [],
    "services": [],
    "tags": [],
    "users": [],         # (email,) — удалить нечем (нет DELETE /users), просто дезактивировать нельзя тоже;
                          # фиксируем для отчёта, реального удаления API не предоставляет.
}

failures: list[str] = []
passed = 0

# Цены тестовых сущностей вынесены в константы: по ним же считаются
# ожидаемые суммы в summary ремонта и в отчётах по услугам.
PART_PURCHASE_PRICE = 10.0
PART_SALE_PRICE = 20.0

SERVICE_A_NAME = f"Wheel true {RUN_ID}"
SERVICE_A_PRICE = 150.0
SERVICE_A_OVERRIDE_PRICE = 250.0   # ручная цена той же услуги во втором списании

SERVICE_B_NAME = f"Brake bleed {RUN_ID}"
SERVICE_B_PRICE = 99.5

SERVICES_TOTAL = SERVICE_A_PRICE + SERVICE_A_OVERRIDE_PRICE + SERVICE_B_PRICE
PARTS_TOTAL = PART_SALE_PRICE * 1


def check(condition: bool, message: str):
    global passed
    if condition:
        passed += 1
        print(f"  ✓ {message}")
    else:
        failures.append(message)
        print(f"  ✗ {message}")


def approx(a, b, tol: float = 0.01) -> bool:
    """Сравнение денежных сумм: цены — float, точного равенства не ждём."""
    return abs(float(a) - float(b)) < tol


def auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login(email: str, password: str) -> str:
    r = client.post("/auth/login", data={"username": email, "password": password})
    if r.status_code != 200:
        raise RuntimeError(f"Login failed for {email}: {r.status_code} {r.text}")
    return r.json()["access_token"]


# ──────────────────────────────────────────────
# Cleanup
# ──────────────────────────────────────────────

def cleanup(admin_headers: dict):
    print("\n=== Cleanup ===")

    for repair_id in created["repairs"]:
        client.delete(f"/repairs/{repair_id}", headers=admin_headers)

    for rental_id in created["rentals"]:
        client.delete(f"/rentals/{rental_id}", headers=admin_headers)

    for person_id, tag_id in created["person_tags"]:
        client.delete(f"/people/{person_id}/tags/{tag_id}", headers=admin_headers)

    for person_id in created["passports"]:
        client.delete(f"/people/{person_id}/passport", headers=admin_headers)

    for person_id in created["people"]:
        client.delete(f"/people/{person_id}", headers=admin_headers)

    for bike_id in created["bikes"]:
        client.delete(f"/bikes/{bike_id}", headers=admin_headers)

    for part_id in created["parts"]:
        client.delete(f"/parts/{part_id}", headers=admin_headers)

    for service_id in created["services"]:
        client.delete(f"/services/{service_id}", headers=admin_headers)

    # У тегов нет отдельного DELETE /tags/{id} по ТЗ (только отвязка от
    # человека), поэтому тестовые теги, созданные этим прогоном, останутся
    # в справочнике — это осознанное ограничение API, а не тест-рантера.
    if created["tags"]:
        print(f"  ⚠ Тестовые теги не удалены (нет DELETE /tags/{{id}} по ТЗ): {created['tags']}")

    if created["users"]:
        print(f"  ⚠ Тестовые пользователи не удалены (нет DELETE /auth/users в ТЗ): {created['users']}")

    print("Cleanup завершён")


# ──────────────────────────────────────────────
# Тесты
# ──────────────────────────────────────────────

def test_auth(admin_h: dict) -> tuple[dict, dict, int, int]:
    print("\n=== Auth / роли (2.7) ===")

    mech_email = f"mech_{RUN_ID}@test.local"
    mgr_email = f"mgr_{RUN_ID}@test.local"

    r = client.post(
        "/auth/register",
        json={"email": mech_email, "full_name": "Test Mechanic", "password": "pass123", "role": "mechanic"},
        headers=admin_h,
    )
    check(r.status_code == 201, "admin может зарегистрировать mechanic")
    created["users"].append(mech_email)

    r = client.post(
        "/auth/register",
        json={"email": mgr_email, "full_name": "Test Manager", "password": "pass123", "role": "manager"},
        headers=admin_h,
    )
    check(r.status_code == 201, "admin может зарегистрировать manager")
    created["users"].append(mgr_email)

    mech_token = login(mech_email, "pass123")
    mgr_token = login(mgr_email, "pass123")
    mech_h = auth_headers(mech_token)
    mgr_h = auth_headers(mgr_token)

    r = client.post(
        "/auth/register",
        json={"email": f"x_{RUN_ID}@test.local", "full_name": "X", "password": "p", "role": "mechanic"},
        headers=mech_h,
    )
    check(r.status_code == 403, "mechanic НЕ может регистрировать пользователей (только admin)")

    r = client.get("/auth/me", headers=mech_h)
    check(r.status_code == 200 and r.json()["role"] == "mechanic", "GET /auth/me возвращает роль mechanic")

    r = client.get("/no-token-should-fail")
    check(r.status_code in (401, 404), "запрос без токена не проходит как защищённый ресурс")

    r = client.get("/", )
    check(r.status_code == 200, "GET / доступен без авторизации")

    return mech_h, mgr_h, None, None


def test_bikes(mech_h: dict, mgr_h: dict) -> int:
    print("\n=== Bikes (write: admin+mechanic) ===")

    r = client.post(
        "/bikes",
        json={"type": "Механический велосипед", "serial_number": f"SN-{RUN_ID}-1"},
        headers=mech_h,
    )
    check(r.status_code == 201, "mechanic создаёт велосипед")
    bike_id = r.json()["id"]
    created["bikes"].append(bike_id)

    r = client.post(
        "/bikes",
        json={"type": "Механический велосипед", "serial_number": f"SN-{RUN_ID}-2"},
        headers=mgr_h,
    )
    check(r.status_code == 403, "manager НЕ может создавать велосипед")

    r = client.get(f"/bikes/{bike_id}", headers=mgr_h)
    check(r.status_code == 200, "manager может читать велосипед (read открыт всем ролям)")

    return bike_id


def test_people(mech_h: dict, mgr_h: dict) -> int:
    print("\n=== People + Tags (2.5, write people: admin+manager) ===")

    r = client.post(
        "/people",
        json={"first_name": "Ivan", "last_name": f"Test{RUN_ID}", "phone": f"+7000{RUN_ID}"},
        headers=mgr_h,
    )
    check(r.status_code == 201, "manager создаёт клиента")
    person_id = r.json()["id"]
    created["people"].append(person_id)
    check(r.json()["tags"] == [], "у нового клиента пустой список тегов")

    r = client.post(
        "/people",
        json={"first_name": "X", "last_name": "Y", "phone": f"+7001{RUN_ID}"},
        headers=mech_h,
    )
    check(r.status_code == 403, "mechanic НЕ может создавать клиента")

    # Теги — доступны всем ролям
    r = client.get("/tags", headers=mech_h)
    check(r.status_code == 200 and len(r.json()) >= 4, "GET /tags возвращает предустановленные теги")
    tags = r.json()
    tag_id = tags[0]["id"]
    tag_name = tags[0]["name"]

    r = client.post(f"/people/{person_id}/tags", json={"tag_id": tag_id}, headers=mech_h)
    check(r.status_code == 201 and tag_name in r.json()["tags"], "mechanic может привязать тег клиенту")
    created["person_tags"].append((person_id, tag_id))

    r = client.get(f"/people?tag={tag_name}", headers=mgr_h)
    check(
        r.status_code == 200 and any(p["id"] == person_id for p in r.json()),
        "GET /people?tag= фильтрует по тегу",
    )

    r = client.delete(f"/people/{person_id}/tags/{tag_id}", headers=mgr_h)
    check(r.status_code == 200 and r.json()["tags"] == [], "тег успешно отвязан от клиента")
    created["person_tags"].remove((person_id, tag_id))

    # Новый тег
    new_tag_name = f"TestTag-{RUN_ID}"
    r = client.post("/tags", json={"name": new_tag_name}, headers=mgr_h)
    check(r.status_code == 201, "manager может создать новый тег")
    created["tags"].append(r.json()["id"])

    # Паспортные данные — write: admin+manager, доступ на чтение всем ролям
    r = client.post(
        f"/people/{person_id}/passport",
        json={"series": "1234", "number": "567890"},
        headers=mgr_h,
    )
    check(r.status_code == 201, "manager создаёт паспортные данные клиента")
    created["passports"].append(person_id)

    r = client.get(f"/people/{person_id}/passport", headers=mech_h)
    check(r.status_code == 200, "mechanic может ПРОСМАТРИВАТЬ паспортные данные (доступ всем ролям)")

    r = client.post(
        f"/people/{person_id}/passport",
        json={"series": "1", "number": "1"},
        headers=mech_h,
    )
    check(r.status_code == 403, "mechanic НЕ может создавать/менять паспортные данные")

    return person_id


def test_parts(mech_h: dict, mgr_h: dict) -> int:
    print("\n=== Parts (2.1 min_stock, write: admin+mechanic) ===")

    r = client.post(
        "/parts",
        json={
            "name": f"Chain-{RUN_ID}",
            "purchase_price": PART_PURCHASE_PRICE,
            "sale_price": PART_SALE_PRICE,
            "owner": "kirill",
            "quantity": 1,
            "min_stock": 2,
        },
        headers=mech_h,
    )
    check(r.status_code == 201, "mechanic создаёт запчасть с min_stock")
    part_id = r.json()["id"]
    check(r.json()["min_stock"] == 2, "min_stock сохранён корректно")
    created["parts"].append(part_id)

    r = client.post(
        "/parts",
        json={"name": "X", "purchase_price": 1, "sale_price": 2, "owner": "vitaly", "quantity": 5},
        headers=mgr_h,
    )
    check(r.status_code == 403, "manager НЕ может создавать запчасти")

    r = client.get("/parts?low_stock=true", headers=mech_h)
    check(
        r.status_code == 200 and any(p["id"] == part_id for p in r.json()),
        "GET /parts?low_stock=true находит позицию с quantity <= min_stock",
    )

    return part_id


def test_parts_merge(mech_h: dict):
    print("\n=== Parts merge (одинаковые позиции объединяются) ===")

    base = {
        "name": f"Камера 29 {RUN_ID}",
        "category": "Расходники",
        "purchase_price": 120,
        "sale_price": 300,
        "owner": "kirill",
        "quantity": 20,
        "min_stock": 2,
    }

    r = client.post("/parts", json=base, headers=mech_h)
    check(r.status_code == 201, "создана позиция «Камера 29» 20 шт по 300")
    first_id = r.json()["id"]
    created["parts"].append(first_id)

    # Та же номенклатура и цены, но название с другим регистром и пробелами.
    r = client.post(
        "/parts",
        json={**base, "quantity": 10, "name": f"  камера 29 {RUN_ID} "},
        headers=mech_h,
    )
    check(
        r.status_code == 201 and r.json()["id"] == first_id and r.json()["quantity"] == 30,
        "повторная поставка той же позиции увеличивает количество (20 + 10 = 30), дубль не создаётся",
    )

    # Пустые справочные поля дозаполняются данными новой поставки.
    r = client.post(
        "/parts",
        json={**base, "quantity": 0, "sku": f"SKU-{RUN_ID}"},
        headers=mech_h,
    )
    check(
        r.status_code == 201 and r.json()["id"] == first_id and r.json()["sku"] == f"SKU-{RUN_ID}",
        "пустой артикул дозаполняется из новой поставки",
    )

    r = client.post("/parts", json={**base, "quantity": 5, "sale_price": 350}, headers=mech_h)
    other_price_id = r.json()["id"]
    check(
        r.status_code == 201 and other_price_id != first_id,
        "позиция с другой ценой продажи не сливается — заводится отдельная строка",
    )
    created["parts"].append(other_price_id)

    r = client.post("/parts", json={**base, "quantity": 5, "owner": "vitaly"}, headers=mech_h)
    other_owner_id = r.json()["id"]
    check(
        r.status_code == 201 and other_owner_id != first_id,
        "позиция другого владельца не сливается — заводится отдельная строка",
    )
    created["parts"].append(other_owner_id)


def test_services(mech_h: dict) -> tuple[int, int]:
    print("\n=== Services (write: admin+mechanic) ===")

    r = client.post(
        "/services",
        json={"name": SERVICE_A_NAME, "price": SERVICE_A_PRICE},
        headers=mech_h,
    )
    check(r.status_code == 201, "mechanic создаёт услугу")
    service_a_id = r.json()["id"]
    created["services"].append(service_a_id)

    r = client.post(
        "/services",
        json={"name": SERVICE_B_NAME, "price": SERVICE_B_PRICE},
        headers=mech_h,
    )
    check(r.status_code == 201, "mechanic создаёт вторую услугу (для проверки сумм в отчётах)")
    service_b_id = r.json()["id"]
    created["services"].append(service_b_id)

    return service_a_id, service_b_id


def test_repairs(
        mech_h: dict,
        mgr_h: dict,
        bike_id: int,
        person_id: int,
        part_id: int,
        service_a_id: int,
        service_b_id: int,
) -> int:
    print("\n=== Repairs (2.2 total_cost, 2.3 waiting_parts, 2.4 closed_by, write: admin+mechanic) ===")

    r = client.post(
        "/repairs",
        json={"bike_id": bike_id, "client_id": person_id, "problem_description": "Скрипит цепь"},
        headers=mech_h,
    )
    check(r.status_code == 201, "mechanic создаёт ремонт")
    repair_id = r.json()["id"]
    created["repairs"].append(repair_id)
    check(r.json()["created_by_user_id"] is not None, "created_by_user_id проставлен автоматически")

    r = client.post(
        "/repairs",
        json={"bike_id": bike_id, "client_id": person_id, "problem_description": "x"},
        headers=mgr_h,
    )
    check(r.status_code == 403, "manager НЕ может создавать ремонт")

    r = client.post(f"/repairs/{repair_id}/parts", json={"part_id": part_id, "quantity": 1}, headers=mech_h)
    check(r.status_code == 201, "mechanic добавляет запчасть в ремонт (со списанием)")

    r = client.post(f"/repairs/{repair_id}/services", json={"service_id": service_a_id}, headers=mech_h)
    check(r.status_code == 201, "mechanic добавляет услугу в ремонт")
    check(
        approx(r.json()["price"], SERVICE_A_PRICE),
        "без явной цены берётся базовая цена услуги из справочника",
    )

    r = client.get(f"/repairs/{repair_id}", headers=mech_h)
    check(
        r.status_code == 200 and approx(r.json()["total_cost"], PART_SALE_PRICE + SERVICE_A_PRICE),
        "total_cost = сумма услуг + запчастей",
    )

    # ── Суммы по услугам: та же услуга второй раз, но с ручной ценой ──
    r = client.post(
        f"/repairs/{repair_id}/services",
        json={"service_id": service_a_id, "price": SERVICE_A_OVERRIDE_PRICE},
        headers=mech_h,
    )
    check(
        r.status_code == 201 and approx(r.json()["price"], SERVICE_A_OVERRIDE_PRICE),
        "ту же услугу можно добавить повторно, ручная цена не подменяется каталожной",
    )

    r = client.post(f"/repairs/{repair_id}/services", json={"service_id": service_b_id}, headers=mech_h)
    check(r.status_code == 201, "вторая услуга добавлена в ремонт")

    r = client.get(f"/repairs/{repair_id}/summary", headers=mech_h)
    summary = r.json()
    check(
        r.status_code == 200 and approx(summary["services_total"], SERVICES_TOTAL),
        f"summary.services_total = {SERVICES_TOTAL} (базовая + ручная + вторая услуга)",
    )
    check(
        approx(summary["parts_total"], PARTS_TOTAL),
        f"summary.parts_total = {PARTS_TOTAL} (списано по цене продажи)",
    )
    check(
        approx(summary["total_for_client"], SERVICES_TOTAL + PARTS_TOTAL),
        "summary.total_for_client = услуги + запчасти",
    )

    r = client.get(f"/repairs/{repair_id}", headers=mech_h)
    check(
        approx(r.json()["total_cost"], SERVICES_TOTAL + PARTS_TOTAL),
        "total_cost в детальном ответе совпадает с summary",
    )

    r = client.get("/repairs", headers=mech_h)
    check(
        r.status_code == 200 and all("total_cost" in item for item in r.json()),
        "GET /repairs (список) содержит total_cost без N+1",
    )

    # 2.3 waiting_parts
    r = client.put(f"/repairs/{repair_id}", json={"status": "waiting_parts"}, headers=mech_h)
    check(r.status_code == 200 and r.json()["status"] == "waiting_parts", "статус waiting_parts принимается")

    r = client.get("/repairs?status=waiting_parts", headers=mech_h)
    check(
        r.status_code == 200 and any(x["id"] == repair_id for x in r.json()),
        "GET /repairs?status=waiting_parts фильтрует корректно",
    )

    # 2.4 closed_by_user_id обязателен при done
    r = client.put(f"/repairs/{repair_id}", json={"status": "done"}, headers=mech_h)
    check(r.status_code == 422, "перевод в done без closed_by_user_id -> 422")

    me = client.get("/auth/me", headers=mech_h).json()
    r = client.put(
        f"/repairs/{repair_id}",
        json={"status": "done", "closed_by_user_id": me["id"]},
        headers=mech_h,
    )
    check(
        r.status_code == 200 and r.json()["closed_by_user_id"] == me["id"] and r.json()["closed_by_name"],
        "перевод в done с closed_by_user_id -> 200, closed_by_name заполнен",
    )

    return repair_id


def test_rentals(mech_h: dict, mgr_h: dict, bike_id: int, person_id: int) -> int:
    print("\n=== Rentals (write: admin+manager) ===")

    # нужен отдельный (свободный) велосипед — repair-тест мог занять/освободить bike_id,
    # но чтобы не зависеть от порядка — создаём новый
    r = client.post(
        "/bikes",
        json={"type": "Механический велосипед", "serial_number": f"SN-{RUN_ID}-rental"},
        headers=mech_h,
    )
    rental_bike_id = r.json()["id"]
    created["bikes"].append(rental_bike_id)

    r = client.post("/rentals", json={"bike_id": rental_bike_id, "person_id": person_id}, headers=mgr_h)
    check(r.status_code == 201, "manager создаёт аренду")
    rental_id = r.json()["id"]
    created["rentals"].append(rental_id)
    check(r.json()["created_by_user_id"] is not None, "created_by_user_id проставлен автоматически")

    r = client.post("/rentals", json={"bike_id": rental_bike_id, "person_id": person_id}, headers=mech_h)
    check(r.status_code == 403, "mechanic НЕ может создавать аренду")

    r = client.post(f"/rentals/{rental_id}/close", json={}, headers=mgr_h)
    check(r.status_code == 200 and r.json()["status"] == "returned", "manager закрывает аренду")

    return rental_id


def test_analytics(mech_h: dict, mgr_h: dict, service_a_id: int, service_b_id: int):
    print("\n=== Analytics (2.6, read-only для всех ролей) ===")

    for path in (
        "/analytics/parts/profit",
        "/analytics/parts/top?limit=5",
        "/analytics/parts/consumption?period=week",
        "/analytics/parts/purchases?period=week",
        "/analytics/services/top?limit=5",
        "/analytics/services/revenue?period=month",
    ):
        r = client.get(path, headers=mech_h)
        check(r.status_code == 200, f"GET {path} -> 200 (mechanic)")
        r = client.get(path, headers=mgr_h)
        check(r.status_code == 200, f"GET {path} -> 200 (manager)")

    # ── Суммы по услугам ─────────────────────────────
    # Проверяем только СВОИ услуги (имена и id из этого прогона) — в БД могут
    # лежать данные прошлых сессий, глобальные итоги на них завязывать нельзя.
    r = client.get("/analytics/services/revenue?period=week", headers=mech_h)
    revenue = r.json()
    items = {item["service_id"]: item for item in revenue["items"]}

    item_a = items.get(service_a_id)
    check(
        item_a is not None
        and item_a["usage_count"] == 2
        and approx(item_a["revenue"], SERVICE_A_PRICE + SERVICE_A_OVERRIDE_PRICE),
        f"services/revenue: услуга A — 2 использования, выручка {SERVICE_A_PRICE + SERVICE_A_OVERRIDE_PRICE} "
        f"(базовая цена + ручная, а не удвоенная каталожная)",
    )

    item_b = items.get(service_b_id)
    check(
        item_b is not None
        and item_b["usage_count"] == 1
        and approx(item_b["revenue"], SERVICE_B_PRICE),
        f"services/revenue: услуга B — 1 использование, выручка {SERVICE_B_PRICE}",
    )

    check(
        approx(revenue["total_revenue"], sum(item["revenue"] for item in revenue["items"])),
        "services/revenue: total_revenue = сумма выручки по позициям",
    )

    r = client.get("/analytics/services/top?limit=100", headers=mech_h)
    top = {item["service_id"]: item for item in r.json()}
    check(
        top.get(service_a_id, {}).get("usage_count") == 2
        and top.get(service_b_id, {}).get("usage_count") == 1,
        "services/top: количество использований совпадает с фактическим",
    )

    r = client.get("/analytics/export?period=month", headers=mgr_h)
    check(
        r.status_code == 200
        and r.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "GET /analytics/export -> 200 + правильный Content-Type xlsx",
    )

    # ── Те же суммы, но внутри выгруженного xlsx ─────
    if load_workbook is None:
        print("  ⚠ openpyxl не установлен — сверка чисел внутри xlsx пропущена")
    elif r.status_code == 200:
        sheet = load_workbook(BytesIO(r.content))["Услуги"]
        rows = {row[0]: row for row in sheet.iter_rows(min_row=2, values_only=True)}

        row_a = rows.get(SERVICE_A_NAME)
        check(
            row_a is not None
            and row_a[1] == 2
            and approx(row_a[2], SERVICE_A_PRICE + SERVICE_A_OVERRIDE_PRICE),
            "xlsx, лист «Услуги»: строка услуги A совпадает с API",
        )

        row_b = rows.get(SERVICE_B_NAME)
        check(
            row_b is not None and row_b[1] == 1 and approx(row_b[2], SERVICE_B_PRICE),
            "xlsx, лист «Услуги»: строка услуги B совпадает с API",
        )


# ──────────────────────────────────────────────
# main
# ──────────────────────────────────────────────

def main():
    print(f"Тестируем {BASE_URL} (run id: {RUN_ID})")

    admin_token = login(ADMIN_EMAIL, ADMIN_PASSWORD)
    admin_h = auth_headers(admin_token)

    try:
        mech_h, mgr_h, _, _ = test_auth(admin_h)

        bike_id = test_bikes(mech_h, mgr_h)
        person_id = test_people(mech_h, mgr_h)
        part_id = test_parts(mech_h, mgr_h)
        test_parts_merge(mech_h)
        service_a_id, service_b_id = test_services(mech_h)

        test_repairs(mech_h, mgr_h, bike_id, person_id, part_id, service_a_id, service_b_id)
        test_rentals(mech_h, mgr_h, bike_id, person_id)
        test_analytics(mech_h, mgr_h, service_a_id, service_b_id)

    finally:
        cleanup(admin_h)

    print(f"\n=== Итого: {passed} passed, {len(failures)} failed ===")
    if failures:
        print("Провалившиеся проверки:")
        for f in failures:
            print(f"  - {f}")
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
