from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi import APIRouter
from fastapi import Depends
from fastapi import Query
from fastapi.responses import StreamingResponse

from sqlalchemy import select
from sqlalchemy import func
from sqlalchemy.ext.asyncio import AsyncSession

from openpyxl import Workbook

from app.db.session import get_db

from app.models.part import Part
from app.models.repair import Repair, RepairPart, RepairService
from app.models.service import Service

from app.enums.owner_type import OwnerType

from app.schemas.analytics import (
    OwnerFinancials,
    PartConsumptionItem,
    PartsConsumptionResponse,
    PartsProfitResponse,
    ServiceRevenueItem,
    ServicesRevenueResponse,
    TopPartItem,
    TopServiceItem,
)


router = APIRouter(
    prefix="/analytics",
    tags=["Analytics"],
)


# ──────────────────────────────────────────────
# Вспомогательные функции
# ──────────────────────────────────────────────

def _period_since(period: str) -> datetime:
    """
    period=week  -> последние 7 дней от текущего момента
    period=month -> последние 30 дней от текущего момента
    """
    days = 7 if period == "week" else 30
    return datetime.now(timezone.utc) - timedelta(days=days)


PeriodQuery = Query(default="week", pattern="^(week|month)$")


# ──────────────────────────────────────────────
# 2.6.1 Запчасти
# ──────────────────────────────────────────────

@router.get(
    "/parts/profit",
    response_model=PartsProfitResponse,
)
async def parts_profit(
        db: AsyncSession = Depends(get_db),
):
    """
    Заработок по запчастям в разбивке по владельцу (Кирилл / Виталий).
    Считается по фактически списанным в ремонты RepairPart
    (а не по остаткам на складе).
    """
    query = (
        select(
            RepairPart.owner,
            func.coalesce(func.sum(RepairPart.purchase_price * RepairPart.quantity), 0).label("cost"),
            func.coalesce(func.sum(RepairPart.sale_price * RepairPart.quantity), 0).label("revenue"),
        )
        .group_by(RepairPart.owner)
    )

    result = await db.execute(query)

    totals = {
        OwnerType.KIRILL: {"cost": 0.0, "revenue": 0.0},
        OwnerType.VITALY: {"cost": 0.0, "revenue": 0.0},
    }

    for owner, cost, revenue in result.all():
        totals[owner] = {"cost": float(cost), "revenue": float(revenue)}

    def _build(owner: OwnerType) -> OwnerFinancials:
        data = totals[owner]
        return OwnerFinancials(
            cost=data["cost"],
            revenue=data["revenue"],
            profit=data["revenue"] - data["cost"],
        )

    return PartsProfitResponse(
        kirill=_build(OwnerType.KIRILL),
        vitaly=_build(OwnerType.VITALY),
    )


@router.get(
    "/parts/top",
    response_model=list[TopPartItem],
)
async def parts_top(
        limit: int = Query(default=10, ge=1, le=100),
        db: AsyncSession = Depends(get_db),
):
    """Топ запчастей по суммарному количеству использований в ремонтах."""
    query = (
        select(
            Part.id,
            Part.name,
            func.coalesce(func.sum(RepairPart.quantity), 0).label("usage_count"),
        )
        .join(RepairPart, RepairPart.part_id == Part.id)
        .group_by(Part.id, Part.name)
        .order_by(func.sum(RepairPart.quantity).desc())
        .limit(limit)
    )

    result = await db.execute(query)

    return [
        TopPartItem(part_id=pid, name=name, usage_count=int(count))
        for pid, name, count in result.all()
    ]


@router.get(
    "/parts/consumption",
    response_model=PartsConsumptionResponse,
)
async def parts_consumption(
        period: str = PeriodQuery,
        db: AsyncSession = Depends(get_db),
):
    """Расход (количество списанных единиц) запчастей за период."""
    since = _period_since(period)

    query = (
        select(
            Part.id,
            Part.name,
            func.coalesce(func.sum(RepairPart.quantity), 0).label("quantity"),
        )
        .join(RepairPart, RepairPart.part_id == Part.id)
        .where(RepairPart.created_at >= since)
        .group_by(Part.id, Part.name)
        .order_by(func.sum(RepairPart.quantity).desc())
    )

    result = await db.execute(query)
    items = [
        PartConsumptionItem(part_id=pid, name=name, quantity=int(qty))
        for pid, name, qty in result.all()
    ]

    return PartsConsumptionResponse(
        period=period,
        since=since.isoformat(),
        total_quantity=sum(i.quantity for i in items),
        items=items,
    )


@router.get(
    "/parts/purchases",
    response_model=PartsConsumptionResponse,
)
async def parts_purchases(
        period: str = PeriodQuery,
        db: AsyncSession = Depends(get_db),
):
    """
    Статистика по 'закупкам' — в проекте нет отдельного учёта закупок
    у поставщика, поэтому это алиас на /parts/consumption (сколько
    продано/списано клиентам через RepairPart за период).
    """
    return await parts_consumption(period=period, db=db)


# ──────────────────────────────────────────────
# 2.6.2 Услуги
# ──────────────────────────────────────────────

@router.get(
    "/services/top",
    response_model=list[TopServiceItem],
)
async def services_top(
        limit: int = Query(default=10, ge=1, le=100),
        db: AsyncSession = Depends(get_db),
):
    """Топ услуг по количеству использований в ремонтах."""
    query = (
        select(
            Service.id,
            Service.name,
            func.count(RepairService.id).label("usage_count"),
        )
        .join(RepairService, RepairService.service_id == Service.id)
        .group_by(Service.id, Service.name)
        .order_by(func.count(RepairService.id).desc())
        .limit(limit)
    )

    result = await db.execute(query)

    return [
        TopServiceItem(service_id=sid, name=name, usage_count=int(count))
        for sid, name, count in result.all()
    ]


@router.get(
    "/services/revenue",
    response_model=ServicesRevenueResponse,
)
async def services_revenue(
        period: str = PeriodQuery,
        db: AsyncSession = Depends(get_db),
):
    """Выручка по услугам за период (последние 7 / 30 дней)."""
    since = _period_since(period)

    query = (
        select(
            Service.id,
            Service.name,
            func.count(RepairService.id).label("usage_count"),
            func.coalesce(func.sum(RepairService.price), 0).label("revenue"),
        )
        .join(RepairService, RepairService.service_id == Service.id)
        .where(RepairService.created_at >= since)
        .group_by(Service.id, Service.name)
        .order_by(func.sum(RepairService.price).desc())
    )

    result = await db.execute(query)
    items = [
        ServiceRevenueItem(
            service_id=sid,
            name=name,
            usage_count=int(count),
            revenue=float(revenue),
        )
        for sid, name, count, revenue in result.all()
    ]

    return ServicesRevenueResponse(
        period=period,
        since=since.isoformat(),
        total_revenue=sum(i.revenue for i in items),
        items=items,
    )


# ──────────────────────────────────────────────
# 2.6.3 Экспорт в Excel
# ──────────────────────────────────────────────

@router.get("/export")
async def export_report(
        period: str = PeriodQuery,
        db: AsyncSession = Depends(get_db),
):
    since = _period_since(period)

    wb = Workbook()

    # ── Лист "Запчасти" ──────────────────────────
    ws_parts = wb.active
    ws_parts.title = "Запчасти"
    ws_parts.append(["Запчасть", "Списано (шт.)", "Выручка", "Маржа"])

    parts_query = (
        select(
            Part.name,
            func.coalesce(func.sum(RepairPart.quantity), 0),
            func.coalesce(func.sum(RepairPart.sale_price * RepairPart.quantity), 0),
            func.coalesce(
                func.sum(
                    (RepairPart.sale_price - RepairPart.purchase_price) * RepairPart.quantity
                ),
                0,
            ),
        )
        .join(RepairPart, RepairPart.part_id == Part.id)
        .where(RepairPart.created_at >= since)
        .group_by(Part.id, Part.name)
        .order_by(Part.name)
    )
    for name, qty, revenue, margin in (await db.execute(parts_query)).all():
        ws_parts.append([name, int(qty), float(revenue), float(margin)])

    # ── Лист "Услуги" ────────────────────────────
    ws_services = wb.create_sheet("Услуги")
    ws_services.append(["Услуга", "Количество", "Выручка"])

    services_query = (
        select(
            Service.name,
            func.count(RepairService.id),
            func.coalesce(func.sum(RepairService.price), 0),
        )
        .join(RepairService, RepairService.service_id == Service.id)
        .where(RepairService.created_at >= since)
        .group_by(Service.id, Service.name)
        .order_by(Service.name)
    )
    for name, count, revenue in (await db.execute(services_query)).all():
        ws_services.append([name, int(count), float(revenue)])

    # ── Лист "Ремонты" ───────────────────────────
    ws_repairs = wb.create_sheet("Ремонты")
    ws_repairs.append(["ID", "Начат", "Статус", "Итого"])

    services_sq = (
        select(
            RepairService.repair_id.label("repair_id"),
            func.coalesce(func.sum(RepairService.price), 0).label("services_sum"),
        )
        .group_by(RepairService.repair_id)
        .subquery()
    )
    parts_sq = (
        select(
            RepairPart.repair_id.label("repair_id"),
            func.coalesce(func.sum(RepairPart.sale_price * RepairPart.quantity), 0).label("parts_sum"),
        )
        .group_by(RepairPart.repair_id)
        .subquery()
    )
    repairs_query = (
        select(
            Repair.id,
            Repair.started_at,
            Repair.status,
            (
                func.coalesce(services_sq.c.services_sum, 0)
                + func.coalesce(parts_sq.c.parts_sum, 0)
            ).label("total_cost"),
        )
        .outerjoin(services_sq, services_sq.c.repair_id == Repair.id)
        .outerjoin(parts_sq, parts_sq.c.repair_id == Repair.id)
        .where(Repair.started_at >= since)
        .order_by(Repair.started_at.desc())
    )
    for repair_id, started_at, status, total in (await db.execute(repairs_query)).all():
        ws_repairs.append([
            repair_id,
            started_at.strftime("%Y-%m-%d %H:%M"),
            status.value if hasattr(status, "value") else status,
            float(total),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    )
